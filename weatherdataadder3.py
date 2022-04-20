
from encodings import utf_8
from openpyxl import *
import pandas as pd
# wb = load_workbook(filename = "D:\\Users\\User\\Desktop\\radon stuff\\radon autodata.xlsx")
# ws = wb.active
# page = wb['RFD']
# uniquedays = []
# for i in range(2, ws.max_row):
#     daydata = ws['E' + str(i)].value
#     daydata = daydata[:10]
#     if daydata not in uniquedays:
#         uniquedays.append(daydata)
# uniquedays = pd.DataFrame(uniquedays)
weathercsv = pd.read_csv("C:\\Users\\Sakhayaan Gavrilyev\\source\\repos\\tensorRadonFlux\\27605.08.11.2018.19.04.2022.1.0.0.ru.utf8.00000000.csv", na_values='', sep =';', skiprows = 6, encoding = 'utf-8', index_col=False)
trashcols = ["P", "Pa", "ff10", "ff3", "N", "W1", "W2", "Tn", "Tx", "Cl", "Nh", "H", "Cm", "Ch", "VV", "tR", "E", "Tg", "E'", "sss"]
weathercsv = weathercsv.drop(trashcols, axis = 1)
weathercsv['RRR'] = weathercsv['RRR'].replace('Осадков нет', 0)
weathercsv['RRR'] = weathercsv['RRR'].replace('Следы осадков', 0)
weathercsv['RRR'] = weathercsv['RRR'].fillna(0)
weathercsv['RRR'] = weathercsv['RRR'].astype(float)
weathercsv['RRR'] = weathercsv['RRR'].groupby(weathercsv.index // 4).transform('sum')
weathercsv['DD'] = weathercsv['DD'].replace('Штиль, безветрие', 'No wind')
weathercsv['DD'] = weathercsv['DD'].replace('Ветер, дующий с севера', 'N')
weathercsv['DD'] = weathercsv['DD'].replace('Ветер, дующий с востока', 'E')
weathercsv['DD'] = weathercsv['DD'].replace('Ветер, дующий с запада', 'W')
weathercsv['DD'] = weathercsv['DD'].replace('Ветер, дующий с юга', 'S')
weathercsv['DD'] = weathercsv['DD'].replace('Ветер, дующий с северо-востока', 'NE')
weathercsv['DD'] = weathercsv['DD'].replace('Ветер, дующий с северо-запада', 'NW')
weathercsv['DD'] = weathercsv['DD'].replace('Ветер, дующий с юго-востока', 'SE')
weathercsv['DD'] = weathercsv['DD'].replace('Ветер, дующий с юго-запада', 'SW')
weathercsv['DD'] = weathercsv['DD'].replace('Ветер, дующий с востоко-северо-востока', 'NEE')
weathercsv['DD'] = weathercsv['DD'].replace('Ветер, дующий с западо-северо-запада', 'NWW')
weathercsv['DD'] = weathercsv['DD'].replace('Ветер, дующий с востоко-юго-востока', 'SEE')
weathercsv['DD'] = weathercsv['DD'].replace('Ветер, дующий с западо-юго-запада', 'SWW')
weathercsv['DD'] = weathercsv['DD'].replace('Ветер, дующий с северо-северо-востока', 'NNE')
weathercsv['DD'] = weathercsv['DD'].replace('Ветер, дующий с северо-северо-запада', 'NNW')
weathercsv['DD'] = weathercsv['DD'].replace('Ветер, дующий с юго-юго-востока', 'SSE')
weathercsv['DD'] = weathercsv['DD'].replace('Ветер, дующий с юго-юго-запада', 'SSW')
weathercsv['WW'] = weathercsv['WW'].replace('Состояние неба в общем не изменилось. ', 'Clear')
weathercsv['WW'] = weathercsv['WW'].replace(' ', 'Clear')
weathercsv['WW'] = weathercsv['WW'].replace('Дымка. ', 'Haze')
weathercsv['WW'] = weathercsv['WW'].replace('Ливневый(ые) дождь(и) слабый(ые) в срок наблюдения или за последний час. ', 'Rain')
weathercsv['WW'] = weathercsv['WW'].replace('Дождь (незамерзающий) неливневый. ', 'Rain')
weathercsv['WW'] = weathercsv['WW'].replace('Дождь незамерзающий непрерывный слабый в срок наблюдения. ', 'Rain')
weathercsv['WW'] = weathercsv['WW'].replace('Дождь незамерзающий с перерывами слабый в срок наблюдения. ', 'Rain')
weathercsv['WW'] = weathercsv['WW'].replace('Ливневый(ые) дождь(и). ', 'Rain')
weathercsv['WW'] = weathercsv['WW'].replace('Ливневый(ые) дождь(и) умеренный(ые) или сильный(ые) в срок наблюдения или за последний час. ', 'Rain')
weathercsv['WW'] = weathercsv['WW'].replace('Снег непрерывный слабый в срок наблюдения. ', 'Snow')
weathercsv['WW'] = weathercsv['WW'].replace('Облака в целом рассеиваются или становятся менее развитыми. ', 'Clear')
weathercsv['WW'] = weathercsv['WW'].replace('Снег с перерывами слабый в срок наблюдения. ', 'Snow')
weathercsv['WW'] = weathercsv['WW'].replace('Снег непрерывный умеренный в срок наблюдения. ', 'Snow')
weathercsv['WW'] = weathercsv['WW'].replace('Снег непрерывный сильный в срок наблюдения. ', 'Snow')
weathercsv['WW'] = weathercsv['WW'].replace('Снег неливневый. ', 'Snow')
weathercsv['WW'] = weathercsv['WW'].replace('Ливневый(ые) дождь(и) со снегом слабый(ые) в срок наблюдения или за последний час. ', 'Snow')
weathercsv['WW'] = weathercsv['WW'].replace('Ливневый снег или ливневый дождь и снег. ', 'Snow')
weathercsv['WW'] = weathercsv['WW'].replace('Дождь со снегом или ледяная крупа неливневые. ', 'Snow')
weathercsv['WW'] = weathercsv['WW'].replace('Ливневый град с дождем или без него, или дождь со снегом без грома слабый в срок наблюдения или за последний час.  Максимальный диаметр градин составляет 3 мм.', 'Snow')
weathercsv['WW'] = weathercsv['WW'].replace('Ливневый снег слабый в срок наблюдения или за последний час. ', 'Snow')
weathercsv['WW'] = weathercsv['WW'].replace('Дождь или морось со снегом слабые. ', 'Drizzle')
weathercsv['WW'] = weathercsv['WW'].replace('Снег с перерывами умеренный в срок наблюдения. ', 'Snow')
weathercsv['WW'] = weathercsv['WW'].replace('Дождь незамерзающий непрерывный умеренный в срок наблюдения. ', 'Rain')
weathercsv['WW'] = weathercsv['WW'].replace('Дождь незамерзающий с перерывами умеренный в срок наблюдения. ', 'Rain')
weathercsv['WW'] = weathercsv['WW'].replace('Морось незамерзающая непрерывная слабая в срок наблюдения. ', 'Drizzle')
weathercsv['WW'] = weathercsv['WW'].replace('Туман или ледяной туман, неба не видно, ослабел за последний час. ', 'Fog')
weathercsv['WW'] = weathercsv['WW'].replace('Туман или ледяной туман, неба не видно, без заметного изменения интенсивности в течение последнего часа. ', 'Fog')
weathercsv['WW'] = weathercsv['WW'].replace('Туман или ледяной туман, неба не видно, начался или усилился в течение последнего часа. ', 'Fog')
weathercsv['WW'] = weathercsv['WW'].replace('Гроза слабая или умеренная без града, но с дождем и/или снегом в срок наблюдения. ', 'Rain')
weathercsv['WW'] = weathercsv['WW'].replace('Ливневый(ые) дождь(и) очень сильный(ые) в срок наблюдения или за последний час. ', 'Rain')
weathercsv['WW'] = weathercsv['WW'].replace('Гроза (с осадками или без них). ', 'Thunder')
weathercsv['WW'] = weathercsv['WW'].replace('Дождь незамерзающий непрерывный сильный в срок наблюдения. ', 'Rain')
weathercsv['WW'] = weathercsv['WW'].replace('Ливневая снежная крупа или небольшой град с дождем или без него, или дождь со снегом слабые в срок наблюдения или за последний час. ', 'Snow')
weathercsv['WW'] = weathercsv['WW'].replace('Морось (незамерзающая) или снежные зерна неливневые. ', 'Drizzle')
weathercsv['WW'] = weathercsv['WW'].replace('Ливневый град, или дождь и град. ', 'Snow')
weathercsv['WW'] = weathercsv['WW'].replace('Морось незамерзающая непрерывная умеренная в срок наблюдения. ', 'Drizzle')
weathercsv['WW'] = weathercsv['WW'].replace('Замерзающая морось или замерзающий дождь неливневые. ', 'Snow')
weathercsv['WW'] = weathercsv['WW'].replace('Дождь замерзающий слабый. ', 'Snow')
weathercsv['WW'] = weathercsv['WW'].replace('Ледяная крупа. ', 'Snow')
weathercsv['WW'] = weathercsv['WW'].replace('Слабый дождь в срок наблюдения. Гроза в течение последнего часа, но не в срок наблюдения. ', 'Rain')
weathercsv['WW'] = weathercsv['WW'].replace('Умеренный или сильный дождь в срок наблюдения. Гроза в течение последнего часа, но не в срок наблюдения. ', 'Rain')
weathercsv['WW'] = weathercsv['WW'].replace('Ливневый снег умеренный или сильный в срок наблюдения или за последний час. ', 'Snow')
weathercsv['WW'] = weathercsv['WW'].replace('Облака в целом образовывались или развивались. ', 'Clear')
weathercsv['WW'] = weathercsv['WW'].replace('Дождь замерзающий умеренный или сильный. ', 'Rain')
weathercsv['WW'] = weathercsv['WW'].replace('Слабый или умеренный поземок в целом низкий (ниже уровня глаз наблюдателя). ', 'Snow')

weathercsv.to_csv('weatherdatacheck.csv')
weathercsv["Местное время в Москве (центр, Балчуг)"] = weathercsv['Местное время в Москве (центр, Балчуг)'].astype(str)

#Filters by time; remove this part and uncomment the output if unfiltered is desired
weather_filt = weathercsv[weathercsv["Местное время в Москве (центр, Балчуг)"].str.contains('15:00')]
weather_filt.to_csv("weather.csv")

#output
#weathercsv.to_csv("weather.csv")
